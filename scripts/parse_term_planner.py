"""
parse_term_planner.py

Builds the term-planner dataset (data/termPlanner.json) from a term sheet of the
MBA batch timetable workbook.

    python3 scripts/parse_term_planner.py "MBA 2025-27 Batch Timetable.xlsx" --term "Term V"

Why this exists separately from lib/parseTimetable.ts:
  parseTimetable.ts feeds Supabase with *this student's* sessions. The planner
  needs the opposite — every section of every course, so it can enumerate valid
  section combinations before the student has enrolled in anything.

Sheet layout:
  Columns A-I : timetable grid
      A - month marker or week marker ("W-1"); merged, carries forward
      B - day of month
      C - day name (Mon, Tue, ...)
      D-I - the six session slots
  Columns K-P : course catalogue
      SL. | Course Title | Code | Credit | Classroom No. | Instructor(s)

Everything the Term V sheet does that Term-IV didn't is handled below and
documented at the point it's handled — see NUMBERS, CODE MATCHING, SESSION
NUMBERING and NON-CLASS TEXT.
"""

import argparse
import json
import re
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

SESSIONS = [
    {"id": 1, "label": "9:00–10:30 am",  "slot_key": "s1"},
    {"id": 2, "label": "10:45–12:15 pm", "slot_key": "s2"},
    {"id": 3, "label": "12:30–2:00 pm",  "slot_key": "s3"},
    {"id": 4, "label": "3:00–4:30 pm",   "slot_key": "s4"},
    {"id": 5, "label": "4:45–6:15 pm",   "slot_key": "s5"},
    {"id": 6, "label": "6:45–8:15 pm",   "slot_key": "s6"},
]

MONTHS = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}

DAY_NAMES = {"mon", "tue", "wed", "thu", "fri", "sat", "sun"}

# ── CODE MATCHING ──────────────────────────────────────────────────────────
# The catalogue and the grid don't spell section codes identically: the
# catalogue has "MSAIC(A)" while the grid writes "MSAIC (A)". Matching is done
# on a canonical form (uppercase, all whitespace stripped) and the display form
# is regenerated consistently, so both sides agree regardless of typing.
#
# ALIASES covers genuine sheet typos where a course is written two ways. "SLM"
# appears for sessions 5-10 of Strategic Leadership, whose catalogue code is
# "SL" and whose sessions 1-4 are written "SL" — the session numbering runs
# continuously across both spellings, and there's a single "SL End Term Exam",
# so they are unambiguously the same course. Left as an explicit, visible map
# rather than fuzzy-matching, because a wrong guess here silently corrupts
# clash detection.
ALIASES = {
    "SLM": "SL",
}


def canon(code: str) -> str:
    """'MSAIC (A)' and 'MSAIC(A)' both -> 'MSAIC(A)'."""
    return re.sub(r"\s+", "", str(code)).upper()


def pretty(code: str) -> str:
    """Display form: normalizes spacing only, never case.

    'MSAIC(A)' -> 'MSAIC (A)', but 'ReM (A)' stays 'ReM (A)'. Case is preserved
    because the sheet's own capitalization (ReM, RuM, ToC) is what students
    recognize — matching is done on canon() instead, so nothing depends on it.
    """
    c = " ".join(str(code).split())
    m = re.match(r"^(.*?)\s*\(\s*([A-Za-z])\s*\)$", c)
    return f"{m.group(1)} ({m.group(2).upper()})" if m else c


# ── SESSION NUMBERING ──────────────────────────────────────────────────────
# Normal: "CRM (A) - S13". The \s* after S covers "CME (A) - S 10" (row 46),
# which is a stray space in the sheet. Not anchored at the end, so trailing
# notes like "IM - S12 => Field Visit" still match cleanly.
ENTRY_RE = re.compile(r"^(.+?)\s*-\s*S\s*(\d+)", re.IGNORECASE)

# "ASSAM - 17" (row 77) — the S is simply missing. Deliberately NOT folded into
# ENTRY_RE: a permissive "- <number>" would happily match "Registration 2:00 -
# 5:00 pm" and invent a class. Only accepted when the prefix is already a known
# catalogue code, and only when the number ends the entry.
BARE_NUM_RE = re.compile(r"^(.+?)\s*-\s*(\d+)\s*$")

# "MG (A) & (B) - S13" — one physical class both sections attend. Two entries,
# so the slot is correctly seen as occupied for either section.
JOINT_RE = re.compile(r"^(.+?)\s*\(([A-Z])\)\s*&\s*\(([A-Z])\)\s*-\s*S\s*(\d+)", re.IGNORECASE)

# ── NON-CLASS TEXT ─────────────────────────────────────────────────────────
NO_CLASS = {"---", "--", "<=>", "", "`", "tbd", "reserved"}

# Exams occupy the day but aren't teaching sessions and don't use the six normal
# slots ("=> 9.30 am" / "=> 2.30 pm"). Collected separately so nothing is lost
# silently, but kept out of `slots` — clash detection is about class timetables.
EXAM_MARKERS = ("end term exam", "mid term exam", "exam tbd", "end term exams")


def num(value):
    """Excel hands back floats for plain integers (13.0, 4.0) — coerce, or None."""
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return int(value)
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return None


def normalize_sheet_name(name: str) -> str:
    """"Term-IV", "Term V", "term_v" all collapse to a comparable form."""
    return re.sub(r"[^a-z0-9]", "", name.lower())


def pick_sheet(wb, target_term: str):
    want = normalize_sheet_name(target_term)
    for name in wb.sheetnames:
        if normalize_sheet_name(name) == want:
            return wb[name], name
    raise SystemExit(
        f'No sheet matching "{target_term}". Available: {", ".join(wb.sheetnames)}'
    )


def parse_month_label(value):
    """Column A month marker -> (label, year, month) or None.

    Term V types these as real dates (datetime(2026, 9, 26) meaning "Sept '26" —
    the day component is noise). Term-IV typed them as text. Both are accepted;
    the year always comes off the value itself, never a hardcoded table, which
    is what made the previous parser need editing every rollover.
    """
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.strftime("%b '%y"), value.year, value.month
    text = str(value).strip()
    if not text or text.lower() in NO_CLASS:
        return None
    m = re.match(r"^([A-Za-z]{3,})[^0-9]*(\d{2,4})?", text)
    if not m:
        return None
    month = MONTHS.get(m.group(1)[:3].lower())
    if month is None:
        return None
    if m.group(2) is None:
        return None  # month with no year is unusable; caller keeps the previous one
    year = int(m.group(2))
    if year < 100:
        year += 2000
    return text, year, month


def parse_courses(ws):
    """Course catalogue from columns K-P. One row can list several section codes."""
    courses = {}          # canonical code -> info
    display = {}          # canonical code -> pretty display code
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=11, max_col=16,
                            values_only=True):
        sl, name, codes, credits, _classroom, instructor = row
        if num(sl) is None or not codes or not name:
            continue
        for line in str(codes).split("\n"):
            raw = line.strip()
            if not raw:
                continue
            key = canon(raw)
            display[key] = pretty(raw)
            courses[key] = {
                "name": str(name).strip(),
                "credits": num(credits) or 0,
                "instructor": " ".join(str(instructor).split()) if instructor else "",
            }
    return courses, display


def parse_entries(cell_text, known, report):
    """One grid cell -> list of {code (canonical), session_num}.

    session_num 0 means the sheet scheduled the class without numbering it
    ("IMC - COIL Interaction", "FT&IT - Additional Session"). Those still occupy
    the slot, so they must count for clash detection.
    """
    if cell_text is None:
        return [], []

    entries, exams = [], []

    # '/' is the sheet's intentional separator; newlines are sometimes a soft
    # wrap and sometimes a second entry, so both are split on — only the code
    # and session number are needed, never the surrounding prose.
    for part in re.split(r"\s*/\s*|\n", str(cell_text)):
        part = " ".join(part.split())
        if not part or part.lower() in NO_CLASS:
            continue

        low = part.lower()
        if any(marker in low for marker in EXAM_MARKERS):
            exams.append(part)
            continue

        joint = JOINT_RE.match(part)
        if joint:
            base, sec1, sec2, n = joint.groups()
            for sec in (sec1, sec2):
                entries.append((canon(f"{base}({sec})"), int(n)))
            continue

        m = ENTRY_RE.match(part)
        if m:
            entries.append((canon(m.group(1)), int(m.group(2))))
            continue

        # Fallbacks below only fire for a prefix that is already a known code,
        # so malformed prose can't invent a class.
        m = BARE_NUM_RE.match(part)
        if m and canon(m.group(1)) in known:
            entries.append((canon(m.group(1)), int(m.group(2))))
            continue

        if " - " in part:
            prefix = part.split(" - ", 1)[0]
            if canon(prefix) in known:
                entries.append((canon(prefix), 0))  # scheduled but unnumbered
                continue

        report.add(part)

    return entries, exams


def parse_timetable(ws, known, report):
    days, exams = [], []
    month_label, year, month = "", None, None
    week = ""

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=10, values_only=True):
        col_a, col_b, col_c = row[0], row[1], row[2]

        if col_a is not None:
            parsed = parse_month_label(col_a)
            if parsed:
                month_label, year, month = parsed
            elif isinstance(col_a, str) and col_a.strip().upper().startswith("W-"):
                week = col_a.strip()

        # A real date row needs a numeric day AND a day name — this is what
        # rejects the header row and the sheet's stray continuation rows.
        day_num = num(col_b)
        if day_num is None:
            continue
        if not isinstance(col_c, str) or col_c.strip()[:3].lower() not in DAY_NAMES:
            continue
        if year is None:
            print(f"  ! day {day_num} skipped — no month resolved yet", file=sys.stderr)
            continue

        iso = f"{year}-{month:02d}-{day_num:02d}"

        slots = {}
        for idx, cell in enumerate(row[3:9], start=1):
            found, found_exams = parse_entries(cell, known, report)
            for raw_code, session_num in found:
                code = ALIASES.get(raw_code, raw_code)
                slots.setdefault(f"s{idx}", []).append(
                    {"code": code, "session_num": session_num}
                )
            for text in found_exams:
                exams.append({"date": iso, "slot": f"s{idx}", "text": text})

        days.append({
            "date": iso,
            "date_num": day_num,
            "month": month_label,
            "week": week,
            "day": col_c.strip(),
            "slots": slots,
        })

    return days, exams


def build(filepath: str, term: str) -> dict:
    wb = load_workbook(filepath, data_only=True)
    ws, sheet_name = pick_sheet(wb, term)
    print(f"  sheet: {sheet_name}")

    catalogue, display = parse_courses(ws)
    known = set(catalogue) | set(ALIASES)
    unmatched = set()
    timetable, exams = parse_timetable(ws, known, unmatched)

    # Rewrite canonical codes to their display form for everything the UI shows.
    for day in timetable:
        for entries in day["slots"].values():
            for entry in entries:
                entry["code"] = display.get(entry["code"], pretty(entry["code"]))

    courses = {display[k]: v for k, v in catalogue.items()}

    subject_sections = {}
    for code, info in courses.items():
        subject_sections.setdefault(info["name"], []).append(code)
    for sections in subject_sections.values():
        sections.sort()

    # A code scheduled in the grid but absent from the catalogue has no credits
    # and no course name, so it can never be selected — it would just silently
    # vanish. Surface it loudly.
    scheduled = {e["code"] for d in timetable for es in d["slots"].values() for e in es}
    orphans = sorted(scheduled - set(courses))
    if orphans:
        print(f"\n  ! {len(orphans)} scheduled code(s) missing from the catalogue:",
              file=sys.stderr)
        for code in orphans:
            print(f"      {code}", file=sys.stderr)

    unscheduled = sorted(set(courses) - scheduled)
    if unscheduled:
        print(f"\n  ! {len(unscheduled)} catalogue code(s) never scheduled:", file=sys.stderr)
        for code in unscheduled:
            print(f"      {code}", file=sys.stderr)

    if unmatched:
        print(f"\n  i {len(unmatched)} cell fragment(s) not read as classes "
              f"(holidays, notices, TBDs — check nothing real is here):", file=sys.stderr)
        for text in sorted(unmatched):
            print(f"      {text}", file=sys.stderr)

    return {
        "term": term,
        "sessions": SESSIONS,
        "courses": courses,
        "subject_sections": subject_sections,
        "timetable": timetable,
        "exams": exams,
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("excel", help="path to the timetable .xlsx")
    ap.add_argument("--term", default="Term V", help='sheet/term name, e.g. "Term V"')
    ap.add_argument("--out", default=None, help="output json path")
    args = ap.parse_args()

    out_path = (Path(args.out) if args.out
                else Path(__file__).resolve().parent.parent / "data" / "termPlanner.json")
    out_path.parent.mkdir(parents=True, exist_ok=True)

    print(f"Parsing {args.excel} ({args.term})")
    data = build(args.excel, args.term)
    out_path.write_text(json.dumps(data, indent=2, ensure_ascii=False))

    scheduled_days = sum(1 for d in data["timetable"] if d["slots"])
    print(f"\nWrote {out_path}")
    print(f"  {len(data['timetable'])} calendar days ({scheduled_days} with classes)")
    print(f"  {len(data['courses'])} sections across {len(data['subject_sections'])} courses")
    print(f"  {len(data['exams'])} exam entries recorded separately")
    if data["timetable"]:
        print(f"  {data['timetable'][0]['date']} → {data['timetable'][-1]['date']}")


if __name__ == "__main__":
    main()
